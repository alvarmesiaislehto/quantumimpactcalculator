"""
Quantum Economic Impact Calculator — Streamlit UI.
Built for IQM Quantum Computers.
Based on McKinsey Quantum Technology Monitor 2026.
"""

import io
import streamlit as st
import plotly.graph_objects as go
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from model import calculate_quantum_impact, calculate_timeline_series
import qri as _qri

# ISO-2 → ISO-3 mapping for the 4 supported countries
_ISO2_TO_ISO3 = {"DE": "DEU", "US": "USA", "FI": "FIN", "IT": "ITA", "GR": "GRC", "PL": "POL", "SI": "SVN", "ES": "ESP", "LU": "LUX", "PT": "PRT"}

@st.cache_data(ttl=3600)
def _get_qri_results():
    iso3_list = list(_ISO2_TO_ISO3.values())
    try:
        return _qri.compute_qri_batch(iso3_list)
    except Exception:
        return {}

# Distinct colours for country traces in multi-country charts
COUNTRY_COLORS = {
    "Germany":       "#C9A84C",
    "United States": "#3A7DC9",
    "Finland":       "#5BBDE0",
    "Italy":         "#E8A020",
    "Greece":        "#4CAF7D",
    "Poland":        "#E05B7D",
    "Slovenia":      "#A78BFA",
    "Spain":         "#F87171",
    "Luxembourg":    "#34D399",
    "Portugal":      "#FB923C",
}

def _hex_to_rgba(hex_color: str, alpha: float = 0.2) -> str:
    h = hex_color.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f"rgba({r},{g},{b},{alpha})"


def _excel_bytes(
    country: str, year: int, scenario: str, apply_readiness: bool,
    sectors_df: pd.DataFrame, total_impact: float, pct_of_gdp: float, full_2035: float,
) -> bytes:
    """Build and return an in-memory Excel workbook as bytes."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    navy_fill  = PatternFill("solid", fgColor="1F3864")
    gold_fill  = PatternFill("solid", fgColor="C9A84C")
    hdr_font   = Font(bold=True, color="FFFFFF")
    val_font   = Font(bold=True, color="1F3864")
    thin       = Side(style="thin", color="CBD5E1")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(ws, row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = navy_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border

    def _val(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = val_font
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border

    meta = [
        ("Country",           country),
        ("Target Year",       year),
        ("Scenario",          scenario),
        ("Readiness Applied", "Yes" if apply_readiness else "No"),
        ("Total Impact ($B)", round(total_impact, 3)),
        ("% of GDP",          f"{pct_of_gdp:.4f}%"),
        ("2035 Full Potential ($B)", round(full_2035, 1)),
    ]
    for i, (label, value) in enumerate(meta, start=2):
        _hdr(ws, i, 1, label)
        _val(ws, i, 2, value)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    # ── Sheet 2: Sector Detail ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Sector Detail")
    headers = [
        "Sector", "Country Sector GDP ($B)", "McKinsey Global QC Value ($B)",
        "World Sector GDP ($B)", "Effective Rate (%)", "Quantum Impact ($B)",
    ]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = navy_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    sorted_df = sectors_df.sort_values("quantum_impact_b", ascending=False)
    for r, (_, row) in enumerate(sorted_df.iterrows(), start=2):
        vals = [
            row["sector"],
            round(row["sector_gdp_b"], 1),
            round(row["qc_global_b"], 0),
            round(row["global_sector_gdp_b"], 0),
            round(row["effective_rate"] * 100, 2),
            round(row["quantum_impact_b"], 3),
        ]
        fill = PatternFill("solid", fgColor="EEF2F7") if r % 2 == 0 else None
        for col, v in enumerate(vals, 1):
            c = ws2.cell(row=r, column=col, value=v)
            c.border = border
            if fill:
                c.fill = fill

    col_widths = [34, 24, 28, 22, 18, 22]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ── Brand colours ────────────────────────────────────────────────────────────
NAVY   = "#1F3864"
GOLD   = "#C9A84C"
NAVY2  = "#2E4D80"   # lighter navy for hover / fill
WHITE  = "#FFFFFF"
DARK_BG    = "#0E1621"   # page background
DARK_CARD  = "#1A2535"   # card / chart background
DARK_GRID  = "#243044"   # subtle gridlines

# Fixed colour per sector — consistent regardless of sort order or year
SECTOR_COLORS = {
    "Chemicals & Materials":         "#1F3864",
    "Financial Services":            "#C9A84C",
    "Pharma & Life Sciences":        "#2E6DA4",
    "Travel, Transport & Logistics": "#D4891A",
    "Advanced Industries":           "#3A7DC9",
    "Energy — Electric Power":       "#E8A020",
    "Energy — Oil & Gas":            "#4A9FE0",
    "Insurance":                     "#B87A15",
    "Telecom":                       "#5BBDE0",
}

# ── Sector descriptions ───────────────────────────────────────────────────────
SECTOR_INFO = {
    "Chemicals & Materials": {
        "summary": (
            "Quantum computers can simulate molecular interactions at the atomic level, "
            "dramatically accelerating the discovery of new materials, catalysts, and "
            "chemical compounds — tasks that are intractable on classical hardware at scale."
        ),
        "use_cases": [
            "Molecular simulation for new catalyst design",
            "Discovery of novel battery and semiconductor materials",
            "Fertiliser and agrochemical process optimisation",
            "Polymer and composite materials discovery",
        ],
        "phase": "Medium adoption 2025–30 · High 2030–35",
        "mckinsey_range": "$450B – $800B",
    },
    "Financial Services": {
        "summary": (
            "Quantum algorithms offer speed advantages in portfolio optimisation, "
            "risk modelling, and options pricing — all large combinatorial problems. "
            "Post-quantum cryptography migration is an immediate, near-term driver."
        ),
        "use_cases": [
            "Portfolio and collateral optimisation",
            "Monte Carlo risk simulation acceleration",
            "Derivatives pricing and options valuation",
            "Post-quantum cryptography (PQC) migration",
        ],
        "phase": "Medium adoption 2025–30 · High 2030–35",
        "mckinsey_range": "$400B – $600B",
    },
    "Pharma & Life Sciences": {
        "summary": (
            "Drug discovery is one of quantum computing's most promising near-term "
            "applications. Accurately simulating how candidate molecules interact with "
            "protein targets requires quantum-scale computation that classical computers "
            "cannot match for large molecules."
        ),
        "use_cases": [
            "Drug molecule and protein interaction simulation",
            "Protein folding and structure prediction",
            "Genomic data analysis and personalised medicine",
            "Clinical trial design optimisation",
        ],
        "phase": "Medium adoption 2025–30 · High 2030–35",
        "mckinsey_range": "$50B – $400B",
    },
    "Travel, Transport & Logistics": {
        "summary": (
            "Route and network scheduling problems are classically NP-hard — quantum "
            "optimisation algorithms can find better solutions faster, cutting fuel costs, "
            "reducing empty miles, and improving on-time performance across complex networks."
        ),
        "use_cases": [
            "Last-mile and long-haul route optimisation",
            "Airline crew and fleet scheduling",
            "Warehouse layout and picking optimisation",
            "Real-time disruption management",
        ],
        "phase": "Medium adoption 2025–30 · High 2030–35",
        "mckinsey_range": "$200B – $500B",
    },
    "Advanced Industries": {
        "summary": (
            "Aerospace, automotive, and electronics manufacturers face enormous "
            "design-space exploration problems. Quantum simulation and optimisation "
            "can reduce R&D cycle times and discover higher-performance components."
        ),
        "use_cases": [
            "Aerodynamic and structural component simulation",
            "Semiconductor process and chip design optimisation",
            "Supply chain scheduling at scale",
            "Robotics path planning and motion optimisation",
        ],
        "phase": "Low–Medium adoption 2025–30 · Medium 2030–35",
        "mckinsey_range": "$200B – $500B",
    },
    "Energy — Electric Power": {
        "summary": (
            "Smart grid management and renewable integration involve large-scale "
            "real-time optimisation problems. Quantum approaches could improve grid "
            "stability, reduce curtailment of renewables, and optimise storage dispatch."
        ),
        "use_cases": [
            "Grid load balancing and demand forecasting",
            "Renewable energy integration and curtailment reduction",
            "Energy storage scheduling and dispatch",
            "Transmission network topology optimisation",
        ],
        "phase": "Low adoption 2025–30 · Medium 2030–35",
        "mckinsey_range": "$50B – $150B",
    },
    "Energy — Oil & Gas": {
        "summary": (
            "Subsurface reservoir modelling and seismic data inversion involve massive "
            "datasets and complex physics simulations. Quantum computers could accelerate "
            "these workflows significantly, reducing exploration risk and drilling costs."
        ),
        "use_cases": [
            "Seismic inversion and reservoir characterisation",
            "Fluid dynamics simulation for reservoir management",
            "Drilling trajectory and well placement optimisation",
            "Upstream logistics and supply chain scheduling",
        ],
        "phase": "Low adoption 2025–30 · Medium 2030–35",
        "mckinsey_range": "$50B – $150B",
    },
    "Insurance": {
        "summary": (
            "Quantum-enhanced Monte Carlo methods and optimisation could speed up "
            "actuarial modelling, improve catastrophe risk pricing, and strengthen "
            "fraud detection — areas where marginal improvements translate to large "
            "financial gains at scale."
        ),
        "use_cases": [
            "Catastrophe risk modelling and pricing",
            "Claims fraud detection and anomaly detection",
            "Regulatory stress testing and capital optimisation",
            "Customer lifetime value and churn prediction",
        ],
        "phase": "Medium adoption 2025–30 · Medium 2030–35",
        "mckinsey_range": "$10B – $50B",
    },
    "Telecom": {
        "summary": (
            "Network topology optimisation and the urgent industry-wide transition to "
            "post-quantum cryptography are the primary quantum use cases for telecoms. "
            "Quantum key distribution (QKD) also offers long-term secure communication."
        ),
        "use_cases": [
            "Network traffic routing and load balancing",
            "5G/6G spectrum allocation optimisation",
            "Post-quantum cryptography (PQC) infrastructure migration",
            "Quantum key distribution (QKD) for secure communications",
        ],
        "phase": "Low adoption 2025–30 · Medium 2030–35",
        "mckinsey_range": "$50B – $100B",
    },
}

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Quantum Economic Impact Calculator",
    page_icon="⚛️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Injected CSS ─────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
  /* ── Dark theme base ── */
  .stApp {{ background-color: #0E1621; }}
  .stApp > header {{ background-color: transparent; }}

  /* Main content area */
  .block-container {{
      background-color: #0E1621;
  }}

  /* Global font + default text colour */
  html, body, [class*="css"] {{
      font-family: "Inter", "Helvetica Neue", Arial, sans-serif;
      color: #CBD5E1;
  }}

  /* Dataframe / table text */
  .stDataFrame, .stDataFrame * {{ color: #CBD5E1 !important; background-color: #1A2535 !important; }}

  /* Expander */
  .streamlit-expanderHeader {{ background-color: #1A2535 !important; color: #CBD5E1 !important; border-radius: 8px; }}
  .streamlit-expanderContent {{ background-color: #1A2535 !important; }}

  /* Spinner */
  .stSpinner > div {{ border-top-color: {GOLD} !important; }}

  /* Sidebar */
  section[data-testid="stSidebar"] {{
      background-color: {NAVY} !important;
  }}
  section[data-testid="stSidebar"] .stMarkdown,
  section[data-testid="stSidebar"] label,
  section[data-testid="stSidebar"] p,
  section[data-testid="stSidebar"] span {{
      color: #DDEEFF !important;
      font-family: "Inter", "Helvetica Neue", Arial, sans-serif;
  }}
  section[data-testid="stSidebar"] h1,
  section[data-testid="stSidebar"] h2,
  section[data-testid="stSidebar"] h3 {{
      color: {GOLD} !important;
  }}
  section[data-testid="stSidebar"] hr {{
      border-color: #3A5A8A;
  }}
  /* Logo — white card */
  section[data-testid="stSidebar"] img {{
      border-radius: 8px;
      padding: 12px 16px;
      background: white;
      display: block;
  }}

  /* Hide sidebar keyboard-shortcut tooltip */
  [data-testid="collapsedControl"] + div [role="tooltip"],
  button[data-testid="baseButton-headerNoPadding"] + div,
  .st-emotion-cache-h5rgaw, .st-emotion-cache-1b0udgb,
  [data-testid="stSidebarCollapseButton"] ~ div[role="tooltip"],
  div[data-testid="stTooltipContent"] {{ display: none !important; }}

  /* Metric cards — dark */
  .metric-card {{
      background: #1A2535;
      border: 1px solid #2A3A52;
      border-radius: 10px;
      padding: 18px 16px 14px;
      text-align: center;
      box-shadow: 0 1px 6px rgba(0,0,0,0.3);
      height: 110px;
      display: flex;
      flex-direction: column;
      justify-content: center;
  }}
  .metric-label {{
      color: #64748B;
      font-size: 11px;
      font-weight: 600;
      text-transform: uppercase;
      letter-spacing: 0.6px;
      margin-bottom: 6px;
  }}
  .metric-value {{
      color: {GOLD};
      font-size: 26px;
      font-weight: 700;
      line-height: 1.1;
  }}
  .metric-value-sm {{
      color: {GOLD};
      font-size: 16px;
      font-weight: 700;
      line-height: 1.2;
  }}
  .metric-sub {{
      color: #475569;
      font-size: 11px;
      margin-top: 4px;
  }}

  /* Section headings */
  .section-heading {{
      color: #E2E8F0;
      font-size: 15px;
      font-weight: 700;
      margin: 20px 0 8px;
      padding-bottom: 6px;
      border-bottom: 2px solid {GOLD};
      display: inline-block;
  }}

  /* Page title / subtitle text */
  h1, h2, h3 {{ color: #E2E8F0 !important; }}
  p {{ color: #94A3B8; }}

  /* Footer */
  .footer-text {{
      color: #475569;
      font-size: 11px;
      margin-top: 36px;
      padding-top: 14px;
      border-top: 1px solid #2A3A52;
      line-height: 1.7;
  }}
</style>
""", unsafe_allow_html=True)

# ── Country list (display name → ISO2) ───────────────────────────────────────
COUNTRIES: dict[str, str] = {
    "Germany":        "DE",
    "United States":  "US",
    "Finland":        "FI",
    "Italy":          "IT",
    "Greece":         "GR",
    "Poland":         "PL",
    "Slovenia":       "SI",
    "Spain":          "ES",
    "Luxembourg":     "LU",
    "Portugal":       "PT",
}

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.image("iqm_logo.jpeg", use_container_width=True)
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    selected_country = st.selectbox(
        "Country",
        list(COUNTRIES.keys()),
        index=0,
    )
    iso2 = COUNTRIES[selected_country]

    selected_year = st.slider(
        "Target Year",
        min_value=2026,
        max_value=2035,
        value=2030,
        step=1,
    )

    scenario = st.radio(
        "Scenario",
        ["Conservative", "Base", "Optimistic"],
        index=1,
        help=(
            "Conservative = McKinsey low penetration estimate\n"
            "Base = arithmetic midpoint of low and high\n"
            "Optimistic = McKinsey high penetration estimate"
        ),
    )

    apply_readiness = st.toggle(
        "Apply Country Readiness",
        value=False,
        help=(
            "Applies a national quantum readiness modifier (0–1) based on "
            "government investment, academic ecosystem, and quantum infrastructure. "
            "Source: McKinsey QT Monitor 2026, pp.42–44. "
            "When OFF, all countries use modifier = 1.0."
        ),
    )

    st.markdown("---")
    st.markdown(
        f"<div style='font-size:11px;color:{GOLD};font-weight:600;margin-bottom:6px;'>"
        "Download Results</div>",
        unsafe_allow_html=True,
    )
    # Placeholder — populated after model runs (below)
    _dl_slot = st.empty()

    st.markdown("---")
    st.markdown(f"""
    <div style="font-size:11px;color:#8AAAC8;line-height:1.65;">
      <span style="color:{GOLD};font-weight:600;">Formula</span><br>
      Impact = Country Sector GDP<br>
      &nbsp;&nbsp;× (McKinsey Global QC Value<br>
      &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;÷ World Sector GDP)<br>
      &nbsp;&nbsp;× Timeline Factor<br>
      &nbsp;&nbsp;× Readiness Score<br><br>
      <span style="color:{GOLD};font-weight:600;">Data sources</span><br>
      QC value at stake: McKinsey QTM 2026<br>
      World sector GDP: World Bank + OECD STAN<br>
      Country GDP: World Bank API
    </div>
    """, unsafe_allow_html=True)

# ── Page header ───────────────────────────────────────────────────────────────
readiness_note = " · Country Readiness Applied" if apply_readiness else ""
st.markdown(f"""
<h1 style="color:{NAVY};font-size:24px;font-weight:800;margin-bottom:2px;">
  Quantum Economic Impact Calculator
</h1>
<p style="color:#6B7A9A;font-size:13px;margin-bottom:20px;">
  {selected_country} &nbsp;·&nbsp; {scenario} scenario &nbsp;·&nbsp; {selected_year}{readiness_note}
</p>
""", unsafe_allow_html=True)

# ── Compute model ─────────────────────────────────────────────────────────────
with st.spinner(f"Fetching World Bank data for {selected_country}…"):
    try:
        sector_results, total_impact, pct_of_gdp = calculate_quantum_impact(
            iso2, selected_year, scenario, apply_readiness
        )
        timeline_df = calculate_timeline_series(iso2, scenario, apply_readiness)
        compute_ok = True
    except Exception as exc:
        st.error(f"Could not compute impact: {exc}")
        compute_ok = False

if not compute_ok:
    st.stop()

sectors_df = (
    pd.DataFrame(sector_results)
    .sort_values("quantum_impact_b", ascending=False)
    .reset_index(drop=True)
)
# Limit display to the three largest sectors by quantum impact
top3_df = sectors_df.head(3).copy()

largest_sector     = sectors_df.iloc[0]["sector"]
largest_sector_val = sectors_df.iloc[0]["quantum_impact_b"]
timeline_factor_used = sectors_df.iloc[0]["timeline_factor"]
full_2035_b = total_impact / timeline_factor_used if timeline_factor_used > 0 else total_impact

# Wire the sidebar download button now that all values are available
_dl_slot.download_button(
    label="⬇ Download Excel",
    data=_excel_bytes(
        selected_country, selected_year, scenario, apply_readiness,
        sectors_df, total_impact, pct_of_gdp, full_2035_b,
    ),
    file_name=f"quantum_impact_{selected_country.replace(' ', '_')}_{selected_year}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

# ── Metric cards ──────────────────────────────────────────────────────────────
c1, c2, c3, c4, c5 = st.columns(5)
cards = [
    (c1, "Total Quantum Impact",
     f"${total_impact:.1f}B",
     f"by {selected_year}",
     False),
    (c2, "2035 Full Potential",
     f"${full_2035_b:.1f}B",
     "at full adoption (2035)",
     False),
    (c3, "% of GDP",
     f"{pct_of_gdp:.4f}%",
     "quantum impact share",
     False),
    (c4, "Largest Sector",
     largest_sector,
     f"${largest_sector_val:.2f}B impact",
     True),
    (c5, "Scenario",
     scenario,
     f"Target year: {selected_year}",
     False),
]
for col, label, value, sub, small in cards:
    val_class = "metric-value-sm" if small else "metric-value"
    col.markdown(
        f'<div class="metric-card">'
        f'<div class="metric-label">{label}</div>'
        f'<div class="{val_class}">{value}</div>'
        f'<div class="metric-sub">{sub}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )

st.markdown("<br>", unsafe_allow_html=True)

# ── Charts: bar + line ────────────────────────────────────────────────────────
col_bar, col_line = st.columns(2)

with col_bar:
    st.markdown(
        f'<div class="section-heading">Top 3 Sectors by Impact — {selected_year}</div>',
        unsafe_allow_html=True,
    )
    bar_df = top3_df.copy()
    colors = [SECTOR_COLORS.get(s, "#1F3864") for s in bar_df["sector"]]

    fig_bar = go.Figure(go.Bar(
        x=bar_df["quantum_impact_b"],
        y=bar_df["sector"],
        orientation="h",
        marker_color=colors,
        text=[f"${v:.2f}B" for v in bar_df["quantum_impact_b"]],
        textposition="outside",
        textfont=dict(size=10.5, color="#CBD5E1"),
        hovertemplate=(
            "<b>%{y}</b><br>"
            "Impact: $%{x:.2f}B<br>"
            "<extra></extra>"
        ),
    ))
    fig_bar.update_layout(
        plot_bgcolor=DARK_CARD,
        paper_bgcolor=DARK_BG,
        margin=dict(l=8, r=70, t=8, b=36),
        height=370,
        font=dict(family="Inter, Helvetica Neue, Arial, sans-serif", size=12, color="#CBD5E1"),
        xaxis=dict(
            title="Quantum Impact ($B)",
            gridcolor=DARK_GRID,
            title_font=dict(size=11),
        ),
        yaxis=dict(autorange="reversed", tickfont=dict(size=11)),
    )
    st.plotly_chart(fig_bar, use_container_width=True)

with col_line:
    st.markdown(
        '<div class="section-heading">Quantum Impact Timeline 2026–2035</div>',
        unsafe_allow_html=True,
    )
    fig_line = go.Figure()

    # Area fill under the curve
    fig_line.add_trace(go.Scatter(
        x=timeline_df["year"],
        y=timeline_df["total_impact_b"],
        mode="lines+markers",
        name="Total Impact ($B)",
        line=dict(color=NAVY, width=2.5),
        marker=dict(
            color=timeline_df["year"].apply(
                lambda y: GOLD if y == selected_year else NAVY2
            ),
            size=timeline_df["year"].apply(
                lambda y: 12 if y == selected_year else 7
            ),
            line=dict(color=NAVY, width=1.5),
        ),
        fill="tozeroy",
        fillcolor="rgba(31,56,100,0.07)",
        hovertemplate=(
            "<b>%{x}</b><br>"
            "Impact: $%{y:.2f}B<br>"
            "<extra></extra>"
        ),
    ))

    # Vertical line at selected year
    sel_val = timeline_df.loc[timeline_df["year"] == selected_year, "total_impact_b"]
    if not sel_val.empty:
        fig_line.add_vline(
            x=selected_year,
            line_dash="dot",
            line_color=GOLD,
            line_width=1.5,
            annotation_text=f"  {selected_year}",
            annotation_font=dict(color=GOLD, size=11),
        )

    fig_line.update_layout(
        plot_bgcolor=DARK_CARD,
        paper_bgcolor=DARK_BG,
        margin=dict(l=8, r=8, t=8, b=36),
        height=370,
        font=dict(family="Inter, Helvetica Neue, Arial, sans-serif", size=12, color="#CBD5E1"),
        xaxis=dict(
            title="Year",
            tickvals=list(range(2026, 2036)),
            tickangle=0,
            gridcolor=DARK_GRID,
            title_font=dict(size=11),
        ),
        yaxis=dict(
            title="Total Quantum Impact ($B)",
            gridcolor=DARK_GRID,
            title_font=dict(size=11),
        ),
        showlegend=False,
    )
    st.plotly_chart(fig_line, use_container_width=True)

# ── Data table ────────────────────────────────────────────────────────────────
st.markdown(
    '<div class="section-heading">Top 3 Sectors — Detail</div>',
    unsafe_allow_html=True,
)

display_df = top3_df[[
    "sector", "sector_gdp_b", "qc_global_b", "global_sector_gdp_b",
    "effective_rate", "quantum_impact_b",
]].copy()

display_df.columns = [
    "Sector",
    "Country Sector GDP ($B)",
    "McKinsey Global QC Value ($B)",
    "World Sector GDP ($B)",
    "Effective Rate",
    "Quantum Impact ($B)",
]

display_df["Country Sector GDP ($B)"]      = display_df["Country Sector GDP ($B)"].map("{:,.1f}".format)
display_df["McKinsey Global QC Value ($B)"]= display_df["McKinsey Global QC Value ($B)"].map("{:,.0f}".format)
display_df["World Sector GDP ($B)"]        = display_df["World Sector GDP ($B)"].map("{:,.0f}".format)
display_df["Effective Rate"]               = display_df["Effective Rate"].map("{:.2%}".format)
display_df["Quantum Impact ($B)"]          = display_df["Quantum Impact ($B)"].map("{:,.3f}".format)

st.dataframe(display_df, use_container_width=True, hide_index=True)

# ── Sector insight cards ───────────────────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
st.markdown(
    '<div class="section-heading">How Quantum Computing Helps Each Sector</div>',
    unsafe_allow_html=True,
)
st.markdown(
    "<p style='color:#64748B;font-size:12px;margin-top:4px;margin-bottom:16px;'>"
    "Showing the top 3 sectors for the selected country and year. "
    "Source: McKinsey Quantum Technology Monitor 2026.</p>",
    unsafe_allow_html=True,
)

for _, row in top3_df.iterrows():
    sector_name = row["sector"]
    info = SECTOR_INFO.get(sector_name, {})
    if not info:
        continue

    color = SECTOR_COLORS.get(sector_name, NAVY)
    impact_val = row["quantum_impact_b"]
    eff_rate   = row["effective_rate"]
    qc_global  = row["qc_global_b"]

    with st.expander(
        f"**{sector_name}** — ${impact_val:.2f}B impact by {selected_year}",
        expanded=True,
    ):
        col_left, col_right = st.columns([3, 1])

        with col_left:
            st.markdown(
                f"<p style='color:#CBD5E1;font-size:13.5px;line-height:1.65;"
                f"margin-bottom:10px;'>{info['summary']}</p>",
                unsafe_allow_html=True,
            )
            bullets = "".join(
                f"<li style='margin-bottom:4px;color:#94A3B8;'>{uc}</li>"
                for uc in info["use_cases"]
            )
            st.markdown(
                f"<b style='color:#E2E8F0;font-size:12px;'>Key quantum use cases</b>"
                f"<ul style='margin-top:6px;padding-left:18px;font-size:13px;"
                f"line-height:1.6;'>{bullets}</ul>",
                unsafe_allow_html=True,
            )

        with col_right:
            st.markdown(
                f"""
                <div style="background:#1A2535;border:1px solid #2A3A52;border-radius:8px;
                            padding:14px 12px;text-align:center;">
                  <div style="color:#64748B;font-size:10px;font-weight:600;
                              text-transform:uppercase;letter-spacing:0.5px;margin-bottom:6px;">
                    McKinsey Global Value at Stake
                  </div>
                  <div style="color:{GOLD};font-size:15px;font-weight:700;">
                    {info['mckinsey_range']}
                  </div>
                  <hr style="border-color:#2A3A52;margin:10px 0;">
                  <div style="color:#64748B;font-size:10px;font-weight:600;
                              text-transform:uppercase;letter-spacing:0.5px;margin-bottom:4px;">
                    Effective Rate
                  </div>
                  <div style="color:{GOLD};font-size:15px;font-weight:700;">
                    {eff_rate:.2%}
                  </div>
                  <hr style="border-color:#2A3A52;margin:10px 0;">
                  <div style="color:#64748B;font-size:10px;font-weight:600;
                              text-transform:uppercase;letter-spacing:0.5px;margin-bottom:4px;">
                    Adoption Phase
                  </div>
                  <div style="color:#94A3B8;font-size:11px;line-height:1.4;">
                    {info['phase']}
                  </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

# ── Country comparison ────────────────────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
st.markdown(
    '<div class="section-heading">Country Comparison</div>',
    unsafe_allow_html=True,
)
st.markdown(
    f"<p style='color:#64748B;font-size:12px;margin-top:4px;margin-bottom:12px;'>"
    f"All four countries — {scenario} scenario · {selected_year} · "
    f"{'Readiness applied' if apply_readiness else 'No readiness modifier'}</p>",
    unsafe_allow_html=True,
)

with st.spinner("Computing comparison across all countries…"):
    _comp_rows = []
    for _cname, _ciso in COUNTRIES.items():
        try:
            _sr, _ti, _pg = calculate_quantum_impact(_ciso, selected_year, scenario, apply_readiness)
            _tf = _sr[0]["timeline_factor"] if _sr else 1.0
            _full = _ti / _tf if _tf > 0 else _ti
            for _s in _sr:
                _comp_rows.append({
                    "country": _cname,
                    "sector":  _s["sector"],
                    "impact":  _s["quantum_impact_b"],
                })
            _comp_rows.append({
                "country": _cname, "sector": "__total__",
                "impact": _ti, "pct_gdp": _pg, "full_2035": _full,
            })
        except Exception:
            pass

_totals = [r for r in _comp_rows if r["sector"] == "__total__"]
_totals_sorted = sorted(_totals, key=lambda x: x["impact"], reverse=True)
_country_order = [r["country"] for r in _totals_sorted]
_sector_rows   = [r for r in _comp_rows if r["sector"] != "__total__"]

# Stacked bar: sector breakdown per country (sorted by total, largest first)
_all_sectors = list(SECTOR_COLORS.keys())
_fig_comp = go.Figure()
for _sec in _all_sectors:
    _x, _y = [], []
    for _cn in _country_order:
        _match = [r["impact"] for r in _sector_rows if r["country"] == _cn and r["sector"] == _sec]
        _x.append(_cn)
        _y.append(_match[0] if _match else 0)
    _fig_comp.add_trace(go.Bar(
        name=_sec, x=_x, y=_y,
        marker_color=SECTOR_COLORS.get(_sec, NAVY),
        hovertemplate="<b>%{x}</b><br>" + _sec + ": $%{y:.2f}B<extra></extra>",
    ))
_fig_comp.update_layout(
    barmode="stack",
    plot_bgcolor=DARK_CARD, paper_bgcolor=DARK_BG,
    margin=dict(l=8, r=8, t=8, b=8),
    height=340,
    font=dict(family="Inter, Helvetica Neue, Arial, sans-serif", size=12, color="#CBD5E1"),
    xaxis=dict(title="", gridcolor=DARK_GRID),
    yaxis=dict(title="Quantum Impact ($B)", gridcolor=DARK_GRID, title_font=dict(size=11)),
    legend=dict(
        orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0,
        font=dict(size=9), bgcolor="rgba(0,0,0,0)",
    ),
    showlegend=True,
)
st.plotly_chart(_fig_comp, use_container_width=True)

# Summary comparison table
_comp_table = pd.DataFrame([
    {
        "Country":               r["country"],
        "Total Impact ($B)":     f"${r['impact']:.2f}B",
        "% of GDP":              f"{r['pct_gdp']:.4f}%",
        "2035 Full Potential":   f"${r['full_2035']:.1f}B",
    }
    for r in _totals_sorted
])
st.dataframe(_comp_table, use_container_width=True, hide_index=True)

# ── Quantum Readiness ─────────────────────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
st.markdown(
    '<div class="section-heading">Quantum Readiness — How Prepared Is Each Country?</div>',
    unsafe_allow_html=True,
)
st.markdown(
    "<p style='color:#64748B;font-size:12px;margin-top:4px;margin-bottom:16px;'>"
    "The Quantum Readiness Index (QRI) scores each country 0–1 across 6 pillars using "
    "public data from OpenAlex, QS Rankings, UNESCO, World Bank, TOP500, and IQM's "
    "State of Quantum report. A score of 1.0 means globally leading; 0.0 means no "
    "quantum ecosystem. When the Country Readiness toggle is enabled, the QRI score "
    "scales the economic impact estimate down to reflect how likely a country is to "
    "realise its theoretical quantum potential.</p>",
    unsafe_allow_html=True,
)

_qri_data = _get_qri_results()
_PILLARS = [
    "Scientific Foundation", "Talent Pipeline", "Government Commitment",
    "Private Ecosystem", "Infrastructure", "Absorptive Capacity",
]
_PILLAR_SHORT = ["Science", "Talent", "Government", "Ecosystem", "Infrastructure", "Absorptive"]

if _qri_data:
    _col_radar, _col_scores = st.columns([3, 2])

    with _col_radar:
        st.markdown(
            "<p style='color:#94A3B8;font-size:12px;font-weight:600;"
            "margin-bottom:8px;'>6-Pillar Breakdown</p>",
            unsafe_allow_html=True,
        )
        _fig_radar = go.Figure()
        for _cname, _ciso2 in COUNTRIES.items():
            _iso3 = _ISO2_TO_ISO3.get(_ciso2)
            if not _iso3 or _iso3 not in _qri_data:
                continue
            _pillar_scores = [
                _qri_data[_iso3]["pillars"].get(p, 0) for p in _PILLARS
            ]
            # Close the polygon
            _pillar_scores_closed = _pillar_scores + [_pillar_scores[0]]
            _labels_closed = _PILLAR_SHORT + [_PILLAR_SHORT[0]]
            _fig_radar.add_trace(go.Scatterpolar(
                r=_pillar_scores_closed,
                theta=_labels_closed,
                fill="toself",
                fillcolor=_hex_to_rgba(COUNTRY_COLORS.get(_cname, NAVY), 0.2),
                line=dict(color=COUNTRY_COLORS.get(_cname, NAVY), width=2),
                name=_cname,
                hovertemplate="<b>" + _cname + "</b><br>%{theta}: %{r:.2f}<extra></extra>",
            ))
        _fig_radar.update_layout(
            polar=dict(
                bgcolor=DARK_CARD,
                radialaxis=dict(
                    visible=True, range=[0, 1],
                    gridcolor=DARK_GRID, tickfont=dict(size=8, color="#64748B"),
                    tickvals=[0.25, 0.5, 0.75, 1.0],
                ),
                angularaxis=dict(
                    tickfont=dict(size=10, color="#CBD5E1"),
                    gridcolor=DARK_GRID,
                ),
            ),
            paper_bgcolor=DARK_BG,
            plot_bgcolor=DARK_BG,
            margin=dict(l=40, r=40, t=20, b=20),
            height=340,
            showlegend=True,
            legend=dict(
                font=dict(size=10, color="#CBD5E1"),
                bgcolor="rgba(0,0,0,0)",
                orientation="h", yanchor="bottom", y=-0.15, xanchor="center", x=0.5,
            ),
            font=dict(family="Inter, Helvetica Neue, Arial, sans-serif"),
        )
        st.plotly_chart(_fig_radar, use_container_width=True)

    with _col_scores:
        st.markdown(
            "<p style='color:#94A3B8;font-size:12px;font-weight:600;"
            "margin-bottom:8px;'>Overall QRI Score</p>",
            unsafe_allow_html=True,
        )
        _score_rows = sorted(
            [
                (cname, _qri_data[_ISO2_TO_ISO3[ciso2]]["qri"])
                for cname, ciso2 in COUNTRIES.items()
                if _ISO2_TO_ISO3.get(ciso2) in _qri_data
            ],
            key=lambda x: x[1], reverse=True,
        )
        _fig_scores = go.Figure(go.Bar(
            x=[s for _, s in _score_rows],
            y=[c for c, _ in _score_rows],
            orientation="h",
            marker_color=[COUNTRY_COLORS.get(c, NAVY) for c, _ in _score_rows],
            text=[f"{s:.3f}" for _, s in _score_rows],
            textposition="outside",
            textfont=dict(size=11, color="#CBD5E1"),
            hovertemplate="<b>%{y}</b><br>QRI: %{x:.3f}<extra></extra>",
        ))
        _fig_scores.update_layout(
            plot_bgcolor=DARK_CARD, paper_bgcolor=DARK_BG,
            margin=dict(l=8, r=60, t=8, b=8),
            height=200,
            xaxis=dict(range=[0, 1.1], gridcolor=DARK_GRID, title="QRI Score (0–1)"),
            yaxis=dict(autorange="reversed"),
            font=dict(family="Inter, Helvetica Neue, Arial, sans-serif", size=11, color="#CBD5E1"),
        )
        st.plotly_chart(_fig_scores, use_container_width=True)

        # Pillar detail table
        _tbl_rows = []
        for _cname, _ciso2 in COUNTRIES.items():
            _iso3 = _ISO2_TO_ISO3.get(_ciso2)
            if not _iso3 or _iso3 not in _qri_data:
                continue
            _row = {"Country": _cname, "QRI": f"{_qri_data[_iso3]['qri']:.3f}"}
            for _p, _ps in zip(_PILLAR_SHORT, _PILLARS):
                _row[_p] = f"{_qri_data[_iso3]['pillars'].get(_ps, 0):.2f}"
            _tbl_rows.append(_row)
        _tbl_rows.sort(key=lambda r: float(r["QRI"]), reverse=True)
        st.dataframe(pd.DataFrame(_tbl_rows), use_container_width=True, hide_index=True)

    # Per-country explanation for the selected country
    _sel_iso3 = _ISO2_TO_ISO3.get(iso2)
    if _sel_iso3 and _sel_iso3 in _qri_data:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(
            f"<p style='color:#94A3B8;font-size:12px;font-weight:600;margin-bottom:6px;'>"
            f"Readiness profile — {selected_country}</p>",
            unsafe_allow_html=True,
        )
        _explanation = _qri.explain_qri(_sel_iso3, _qri_data)
        st.markdown(
            f"<p style='color:#CBD5E1;font-size:13px;line-height:1.7;"
            f"background:#1A2535;border-left:3px solid {GOLD};"
            f"padding:12px 16px;border-radius:4px;'>{_explanation}</p>",
            unsafe_allow_html=True,
        )
else:
    st.info(
        "QRI scores could not be loaded. Run the app with an active internet connection "
        "or ensure qri_cache.json is present."
    )

# ── Debug / validation expander ───────────────────────────────────────────────
VALIDATION_COUNTRIES = ["DE", "US", "FI", "IT", "GR", "PL", "SI", "ES", "LU", "PT"]

with st.expander("🔬 Model Validation (debug)"):
    st.markdown(
        "Runs the model for **Germany, United States, Finland, and Italy** at "
        "**2035, Base scenario, no readiness modifier**. "
        "McKinsey's global quantum value-at-stake range is **$1.28T – $2.7T** (p.17), "
        "calibrated against sector revenues/gross output. "
        "This model applies rates to sector **GDP (value-added)**, which is typically "
        "3–5× smaller than gross output for manufacturing sectors — the country totals "
        "below are a conservative, GDP-based calibration and can be verified against "
        "the methodology PDF."
    )
    if st.button("Run Validation"):
        global_total = 0.0
        val_rows: list[dict] = []
        prog = st.progress(0, text="Fetching data…")
        for i, country_iso in enumerate(VALIDATION_COUNTRIES):
            try:
                _, impact, _ = calculate_quantum_impact(country_iso, 2035, "Base", False)
                global_total += impact
                val_rows.append({"Country": country_iso, "Impact 2035 ($B)": round(impact, 1)})
            except Exception as e:
                val_rows.append({"Country": country_iso, "Impact 2035 ($B)": f"Error: {e}"})
            prog.progress(
                (i + 1) / len(VALIDATION_COUNTRIES),
                text=f"Processing {country_iso}…",
            )
        prog.empty()

        col_v1, col_v2 = st.columns([1, 2])
        with col_v1:
            st.metric(
                "Combined Impact — 4 Countries (2035 Base)",
                f"${global_total:.1f}B",
            )
            st.info(
                "These figures are computed from live World Bank GDP data using "
                "McKinsey penetration rates. See the Methodology PDF for full "
                "per-country breakdowns and verification links."
            )
        with col_v2:
            st.dataframe(
                pd.DataFrame(val_rows),
                use_container_width=True,
                hide_index=True,
            )

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="footer-text">
  Based on <b>McKinsey Quantum Technology Monitor 2026</b>.
  Penetration rates are McKinsey estimates (pp.17–26); sector GDP splits use live
  World Bank API data with OECD STAN crosswalk approximations.
  Country readiness scores derived from a 6-pillar Quantum Readiness Index (QRI)
  using OpenAlex, QS Rankings, UNESCO UIS, World Bank, and TOP500 data.
  Charts show the <b>3 highest-impact sectors</b> per country.
  This tool is for illustrative purposes only and does not constitute investment or commercial advice.
  &nbsp;|&nbsp; Built for <b>IQM Quantum Computers</b>.
</div>
""", unsafe_allow_html=True)
